"""
routes/contabil.py — Funcionalidades contábeis avançadas.

Acesso restrito: só usuários com modo_contabil=1 podem usar.
O administrador ativa isso diretamente no banco por usuário.

Funcionalidades:
- Exportar transações para Excel (.xlsx)
- Lançamento com partida dobrada (débito/crédito)
- Relatório de partida dobrada em Excel
"""

import io
import logging
from datetime import date, datetime

from flask import (
    Blueprint, Response, flash,
    redirect, render_template, request, url_for, abort
)
from flask_login import login_required, current_user

from routes.helpers import get_services
from utils.formatters import parse_valor_monetario

logger = logging.getLogger(__name__)

contabil_bp = Blueprint("contabil", __name__, url_prefix="/contabil")


def requer_modo_contabil(f):
    """Decorator — bloqueia acesso se usuário não tem modo contábil ativo."""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        svc = get_services()
        with svc.db.get_conn() as conn:
            row = conn.execute(
                "SELECT modo_contabil FROM usuarios WHERE id=?",
                (current_user.id,)
            ).fetchone()
        if not row or not row["modo_contabil"]:
            abort(403)
        return f(*args, **kwargs)
    return decorated


# ── Exportar Excel — disponível para TODOS os usuários ────────────────────────

@contabil_bp.route("/exportar")
@login_required
def exportar():
    """Página de exportação de transações para Excel."""
    hoje = date.today()
    inicio_padrao = date(hoje.year, 1, 1).strftime("%Y-%m-%d")
    fim_padrao    = hoje.strftime("%Y-%m-%d")

    svc = get_services()
    with svc.db.get_conn() as conn:
        row = conn.execute(
            "SELECT modo_contabil FROM usuarios WHERE id=?",
            (current_user.id,)
        ).fetchone()
    tem_contabil = bool(row and row["modo_contabil"])

    return render_template(
        "contabil/exportar.html",
        filtro_inicio=request.args.get("inicio", inicio_padrao),
        filtro_fim=request.args.get("fim", fim_padrao),
        tem_contabil=tem_contabil,
    )


@contabil_bp.route("/exportar/download")
@login_required
def exportar_download():
    """Gera e baixa o arquivo Excel com as transações do período."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("Biblioteca openpyxl não instalada. Rode: pip install openpyxl", "erro")
        return redirect(url_for("contabil.exportar"))

    svc = get_services()
    uid = current_user.id
    hoje = date.today()

    inicio = request.args.get("inicio", date(hoje.year, 1, 1).strftime("%Y-%m-%d"))
    fim    = request.args.get("fim",    hoje.strftime("%Y-%m-%d"))

    transacoes = svc.transacoes.listar_por_periodo(inicio, fim, uid)

    # ── Cria planilha ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transações"

    # Cores
    COR_HEADER  = "2D1B69"  # roxo escuro
    COR_RECEITA = "D4F5E6"  # verde claro
    COR_DESPESA = "FFE4E4"  # vermelho claro
    COR_TITULO  = "F0EBFF"  # roxo claro

    # ── Título ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    titulo = ws["A1"]
    titulo.value = f"Gravs — Extrato Financeiro | {inicio} até {fim}"
    titulo.font = Font(name="Calibri", size=13, bold=True, color="2D1B69")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    titulo.fill = PatternFill("solid", fgColor=COR_TITULO)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value = f"Usuário: {current_user.nome} | Exportado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sub.font = Font(name="Calibri", size=9, color="666666")
    sub.alignment = Alignment(horizontal="center")

    # ── Cabeçalhos ─────────────────────────────────────────────────────────────
    cabecalhos = ["Data", "Descrição", "Categoria", "Tipo", "Conta/Cartão", "Valor (R$)", "Parcelado", "Recorrente"]
    for col, cab in enumerate(cabecalhos, 1):
        cell = ws.cell(row=4, column=col, value=cab)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[4].height = 22

    # ── Dados ──────────────────────────────────────────────────────────────────
    total_receitas = 0.0
    total_despesas = 0.0

    for linha, t in enumerate(transacoes, 5):
        is_receita = t["tipo"] == "receita"
        cor = COR_RECEITA if is_receita else COR_DESPESA
        fill = PatternFill("solid", fgColor=cor)

        # Formata conta/cartão para exibição
        conta_txt = ""
        if t.get("conta_icone") and t.get("conta_nome"):
            conta_txt = f"{t['conta_icone']} {t['conta_nome']}"
        elif t.get("conta_nome"):
            conta_txt = t["conta_nome"]

        dados = [
            t["data"],
            t["descricao"],
            f"{t.get('categoria_icone', '')} {t.get('categoria_nome', '')}".strip(),
            "Receita" if is_receita else "Despesa",
            conta_txt,
            t["valor"],
            "Sim" if t.get("grupo_parcela") else "Não",
            "Sim" if t.get("recorrente_uuid") else "Não",
        ]

        for col, valor in enumerate(dados, 1):
            cell = ws.cell(row=linha, column=col, value=valor)
            cell.fill = fill
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center")

            if col == 5:  # Valor
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(
                    name="Calibri", size=10, bold=True,
                    color="1A7A4A" if is_receita else "C0392B"
                )

        if is_receita:
            total_receitas += t["valor"]
        else:
            total_despesas += t["valor"]

    # ── Totais ─────────────────────────────────────────────────────────────────
    linha_total = len(transacoes) + 6
    ws.cell(row=linha_total, column=1, value="").fill = PatternFill("solid", fgColor="EEEEEE")

    resumo = [
        ("Total Receitas", total_receitas, "1A7A4A"),
        ("Total Despesas", total_despesas, "C0392B"),
        ("Saldo do Período", total_receitas - total_despesas,
         "1A7A4A" if total_receitas >= total_despesas else "C0392B"),
    ]

    for i, (label, valor, cor_txt) in enumerate(resumo):
        row = linha_total + i
        ws.merge_cells(f"A{row}:D{row}")
        cell_label = ws.cell(row=row, column=1, value=label)
        cell_label.font = Font(name="Calibri", size=10, bold=True)
        cell_label.alignment = Alignment(horizontal="right")
        cell_label.fill = PatternFill("solid", fgColor="F5F5F5")

        cell_valor = ws.cell(row=row, column=5, value=valor)
        cell_valor.number_format = 'R$ #,##0.00'
        cell_valor.font = Font(name="Calibri", size=11, bold=True, color=cor_txt)
        cell_valor.alignment = Alignment(horizontal="right")
        cell_valor.fill = PatternFill("solid", fgColor="F5F5F5")

    # ── Largura das colunas ────────────────────────────────────────────────────
    larguras = [12, 35, 18, 12, 20, 16, 12, 12]
    for col, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = larg

    # ── Salva em memória e retorna ─────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_arquivo = f"gravs_extrato_{inicio}_{fim}.xlsx"
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"}
    )


# ── Partida Dobrada — só para usuários com modo_contabil=1 ────────────────────

@contabil_bp.route("/partida-dobrada")
@login_required
@requer_modo_contabil
def partida_dobrada():
    """Lista lançamentos em partida dobrada."""
    svc = get_services()
    uid = current_user.id
    hoje = date.today()

    inicio = request.args.get("inicio", date(hoje.year, hoje.month, 1).strftime("%Y-%m-%d"))
    fim    = request.args.get("fim",    hoje.strftime("%Y-%m-%d"))

    transacoes = svc.transacoes.listar_por_periodo(inicio, fim, uid)
    # Filtra só as que têm partida dobrada
    com_partida = [t for t in transacoes if t.get("conta_debito") or t.get("conta_credito")]

    return render_template(
        "contabil/partida_dobrada.html",
        transacoes=com_partida,
        filtro_inicio=inicio,
        filtro_fim=fim,
    )


@contabil_bp.route("/partida-dobrada/novo", methods=["GET", "POST"])
@login_required
@requer_modo_contabil
def novo_lancamento_contabil():
    """Cria lançamento com partida dobrada."""
    svc = get_services()

    if request.method == "POST":
        try:
            valor = parse_valor_monetario(request.form.get("valor", "0"))

            # Cria a transação normal
            id_t, erros = svc.transacoes.adicionar(
                descricao=request.form.get("descricao", ""),
                valor=valor,
                tipo=request.form.get("tipo", "despesa"),
                categoria_id=int(request.form.get("categoria_id", 0)),
                usuario_id=current_user.id,
                data=request.form.get("data"),
            )

            if not erros and id_t:
                # Adiciona as contas de débito e crédito
                conta_debito  = request.form.get("conta_debito", "").strip()
                conta_credito = request.form.get("conta_credito", "").strip()

                if conta_debito or conta_credito:
                    with svc.transacoes._repo._db.get_write_conn() as conn:
                        conn.execute(
                            "UPDATE transacoes SET conta_debito=?, conta_credito=? WHERE id=?",
                            (conta_debito, conta_credito, id_t)
                        )

                flash("✓ Lançamento contábil registrado!", "sucesso")
            else:
                flash("Erro ao registrar lançamento.", "erro")

        except (ValueError, TypeError) as exc:
            logger.error("Erro em lançamento contábil: %s", exc)
            flash("Erro ao processar. Verifique os dados.", "erro")

        return redirect(url_for("contabil.partida_dobrada"))

    categorias = svc.categorias_repo.listar_por_usuario(current_user.id)
    return render_template(
        "contabil/novo_lancamento.html",
        categorias=categorias,
        data_hoje=date.today().strftime("%Y-%m-%d"),
    )


@contabil_bp.route("/partida-dobrada/exportar")
@login_required
@requer_modo_contabil
def exportar_partida_dobrada():
    """Exporta lançamentos de partida dobrada para Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("Biblioteca openpyxl não instalada.", "erro")
        return redirect(url_for("contabil.partida_dobrada"))

    svc = get_services()
    uid = current_user.id
    hoje = date.today()

    inicio = request.args.get("inicio", date(hoje.year, 1, 1).strftime("%Y-%m-%d"))
    fim    = request.args.get("fim",    hoje.strftime("%Y-%m-%d"))

    transacoes = svc.transacoes.listar_por_periodo(inicio, fim, uid)
    com_partida = [t for t in transacoes if t.get("conta_debito") or t.get("conta_credito")]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Partida Dobrada"

    # Cabeçalho
    ws.merge_cells("A1:G1")
    ws["A1"].value = f"Gravs — Lançamentos em Partida Dobrada | {inicio} a {fim}"
    ws["A1"].font = Font(bold=True, size=13, color="2D1B69")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = PatternFill("solid", fgColor="F0EBFF")

    cabecalhos = ["Data", "Descrição", "Tipo", "Valor (R$)", "Conta Débito", "Conta Crédito", "Categoria"]
    for col, cab in enumerate(cabecalhos, 1):
        cell = ws.cell(row=3, column=col, value=cab)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2D1B69")
        cell.alignment = Alignment(horizontal="center")

    for linha, t in enumerate(com_partida, 4):
        ws.cell(row=linha, column=1, value=t["data"])
        ws.cell(row=linha, column=2, value=t["descricao"])
        ws.cell(row=linha, column=3, value="Receita" if t["tipo"] == "receita" else "Despesa")
        val = ws.cell(row=linha, column=4, value=t["valor"])
        val.number_format = 'R$ #,##0.00'
        ws.cell(row=linha, column=5, value=t.get("conta_debito", ""))
        ws.cell(row=linha, column=6, value=t.get("conta_credito", ""))
        ws.cell(row=linha, column=7, value=t.get("categoria_nome", ""))

    larguras = [12, 35, 12, 16, 25, 25, 20]
    for col, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = larg

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=gravs_partida_dobrada_{inicio}_{fim}.xlsx"}
    )
